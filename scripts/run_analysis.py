"""run_analysis.py
=================
Full pipeline: TradingView CSV → Excel analysis workbook.

Usage
-----
    python scripts/run_analysis.py data/raw/YOUR_FILE.csv
    python scripts/run_analysis.py data/raw/YOUR_FILE.csv --signals kern

Signal groups
-------------
    lor   – Lorentzian Buy/Sell flip-cycle signals (default)
    kern  – Kernel confident dot entries in LOR direction

Output
------
    data/out/<filename>_<signals>_<timestamp>.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from renko_system.instruments import INSTRUMENT_CONFIG, detect_instrument_from_filename
from renko_system.preprocess import preprocess_tv_export
from renko_system.trades_lor import mark_lor_trades_into_df
from renko_system.trades_kern import mark_kern_trades_into_df
from renko_system.exits import (
    add_mae_mfe,
    build_exit_suite_from_cal,
    run_exit_strategies,
    trade_core_calibration,
)

MAX_FWD_BARS = 300

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")


def build_baseline_strategies(tick_size: float) -> dict:
    """Fixed exit strategies expressed as tick multiples — no data snooping."""
    t = tick_size
    return {
        "BASELINE":    {},
        "SL10_TP20":   dict(stop_loss=t*10, take_profit=t*20),
        "SL20_TP40":   dict(stop_loss=t*20, take_profit=t*40),
        "SL10_TP40":   dict(stop_loss=t*10, take_profit=t*40),
        "TRAIL_10_8":  dict(stop_loss=t*10, trail_start=t*10, trail_dist=t*8),
        "TRAIL_20_15": dict(stop_loss=t*20, trail_start=t*20, trail_dist=t*15),
        "BE_10_TRAIL": dict(stop_loss=t*10, breakeven_at=t*10, breakeven_dist=0,
                            trail_start=t*15, trail_dist=t*8),
    }


def generate_trades(df: pd.DataFrame, signals: str) -> pd.DataFrame:
    if signals == "lor":
        _, trades = mark_lor_trades_into_df(
            df,
            price_col    = "open",
            lor_buy_col  = "lor_buy_sig",
            lor_sell_col = "lor_sell_sig",
            kern_col     = "kern_est",
            state_col    = "state_code",
        )
    elif signals == "kern":
        _, trades = mark_kern_trades_into_df(
            df,
            buy_col   = "lor_buy_sig",
            sell_col  = "lor_sell_sig",
            kern_col  = "kern_est",
            green_col = "kern_confident",
            red_col   = "kern_not_confident",
            open_col  = "open",
            state_col = "state_code",
        )
    else:
        raise ValueError(f"Unknown signal group: '{signals}'. Choose lor or kern.")
    return trades


def _write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    for col_i, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_i, row in enumerate(df.itertuples(index=False), 2):
        for col_i, val in enumerate(row, 1):
            if isinstance(val, np.integer):                val = int(val)
            elif isinstance(val, np.floating):             val = None if np.isnan(val) else float(val)
            elif isinstance(val, float) and np.isnan(val): val = None
            elif isinstance(val, pd.Timestamp):            val = str(val)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = BODY_FONT
            if row_i % 2 == 0:
                cell.fill = ALT_FILL
    for col_i, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 8)
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50),
                                min_col=col_i, max_col=col_i):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _equity_df(trades_all: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for strat in sorted(trades_all["strategy"].unique()):
        sub = (trades_all[trades_all["strategy"] == strat]
               .sort_values("entry_dt").reset_index(drop=True))
        sub["cum_pnl"] = sub["pnl"].cumsum()
        frames.append(sub[["entry_dt", "cum_pnl"]]
                      .rename(columns={"cum_pnl": strat})
                      .set_index("entry_dt"))
    return pd.concat(frames, axis=1).reset_index() if frames else pd.DataFrame()


def run(csv_path: Path, out_dir: Path, signals: str) -> Path:
    print(f"Loading  : {csv_path.name}")
    print(f"Signals  : {signals.upper()}")

    raw = pd.read_csv(csv_path)
    df  = preprocess_tv_export(raw)

    instrument = detect_instrument_from_filename(csv_path.name)
    tick_size  = INSTRUMENT_CONFIG[instrument]["tick_size"]
    print(f"Instrument: {instrument}  tick={tick_size}  rows: {len(df)}")

    trades = generate_trades(df, signals)
    print(f"Trades   : {len(trades)}")

    if trades.empty:
        print("No trades found — check signal columns in CSV.")
        sys.exit(0)

    trades = add_mae_mfe(trades, df)
    trades["entry_dt"] = df.loc[trades["entry_bar_i"].astype(int), "dt"].dt.strftime("%Y-%m-%d %H:%M:%S").values
    trades["time_bucket"] = df.loc[trades["entry_bar_i"].astype(int), "time_bucket"].values

    # Baseline: tick-scaled, no data snooping
    baseline = build_baseline_strategies(tick_size)

    # CAL_ prefix makes it clear these are calibrated from current data (exploratory)
    cal   = trade_core_calibration(trades, tick_size=tick_size)
    suite = build_exit_suite_from_cal(cal, tick_size=tick_size, prefix="CAL_")
    all_strategies = {**baseline, **{k: v for k, v in suite.items() if not k.startswith("_")}}

    trades_all, summary = run_exit_strategies(
        df, trades, all_strategies, max_fwd_bars=MAX_FWD_BARS
    )
    print(f"Strategies: {len(all_strategies)}  result rows: {len(trades_all)}")

    by_time_rows = []
    for strat in trades_all["strategy"].unique():
        sub = trades_all[trades_all["strategy"] == strat]
        for tb in sub["time_bucket"].unique():
            g  = sub[sub["time_bucket"] == tb]
            eq = g["pnl"].cumsum()
            dd = float((eq - eq.cummax()).min())
            by_time_rows.append({
                "strategy":    strat,
                "time_bucket": tb,
                "trades":      len(g),
                "pnl_sum":     round(g["pnl"].sum(), 4),
                "pnl_mean":    round(g["pnl"].mean(), 4),
                "win_rate":    round((g["pnl"] >= 0).mean(), 4),
                "max_dd":      round(dd, 4),
            })
    by_time = pd.DataFrame(by_time_rows).sort_values(["strategy", "time_bucket"])

    trade_cols = ["trade_id", "side", "entry_dt", "entry_px", "exit_px", "pnl",
                  "bars_held", "mae", "mfe", "state_sig", "time_bucket"]
    trade_cols = [c for c in trade_cols if c in trades.columns]
    trades_out = trades[trade_cols].copy()
    trades_out["pnl_cumsum"] = trades_out["pnl"].cumsum()

    summary_out = summary[["strategy", "time_bucket", "trades", "pnl_sum",
                            "pnl_mean", "win_rate", "max_dd", "pnl_to_dd"]].copy()
    for col in ["pnl_sum", "pnl_mean", "win_rate", "pnl_to_dd"]:
        summary_out[col] = summary_out[col].round(4)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "trades",       trades_out)
    _write_sheet(wb, "exit_summary", summary_out)
    _write_sheet(wb, "by_time",      by_time)
    eq_df = _equity_df(trades_all)
    if not eq_df.empty:
        _write_sheet(wb, "equity", eq_df)

    out_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_dir / f"{csv_path.stem}_{signals}_{ts}.xlsx"
    wb.save(out_path)
    print(f"Saved    : {out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Run trade analysis on a TradingView CSV.")
    ap.add_argument("csv",       type=Path, help="Path to TradingView export CSV")
    ap.add_argument("--signals", type=str,  default="lor", choices=["lor", "kern"],
                    help="Signal group to analyse (default: lor)")
    ap.add_argument("--out",     type=Path, default=Path("data/out"))
    args = ap.parse_args()

    if not args.csv.exists():
        sys.exit(f"File not found: {args.csv}")

    run(args.csv, args.out, args.signals)


if __name__ == "__main__":
    main()
